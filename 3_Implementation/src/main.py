# Author: Karisma Kiran(99003760)
# Contact: karisma.kiran@ltts.com /karismakiran07@gmail.com
# Date of creation: 22/3/2021

# -------------------------------------------------------------------------------------------------------------#
# -------------------------------------------------------------------------------------------------------------#

"""
This program performs the task of putting all the data of a particular candidate from 5 sheets
to a master sheet provided the user gives the input of name , ps no or email id of that particular candidate.
Then, MasterSheet file will give bar charts outputs of any two columns/rows combination.
This programs uses both pandas and openpyxl library and it has object oriented programming concepts such as class,
object and function.
"""

# -------------------------------------------------------------------------------------------------------------#
# -------------------------------------------------------------------------------------------------------------#

"""
This program uses pandas and openpyxl library and import pandas as pd  
imports the library and rather than using the name pandas , it's instructed to use the name pd instead.
From Pandas ExcelWriter is imported in order to write the header in the first master sheet.
"""


# IMPORT

import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference

WORKSHEET = "PythonSheets.xlsx"
SHEETS = ["Sheet1", "Sheet2", "Sheet3", "Sheet4", "Sheet5"]
MSHEET = "MasterSheet"


"""
This program uses a class named Aggregator where the __init__ function is defined and has three 
arguments such as self,worksheet, and sheets. The worksheet has 5 sheets and for each sheet the
specific function is defined in this section. The __init__ method is called and then we're reading the excel file.
"""

class Aggregator:
    def __init__(self, worksheet, sheets):
        self.worksheet, self.sheets = worksheet, sheets
        self.dfs = pd.read_excel(worksheet, sheet_name=sheets, parse_dates=False)

    #  Validation (PS Number) from all sheets

    def get_input(self, c=0):
        query = input("Enter PS Number / Email / Name: ")        # query is what we've provided in input
        if query and c < 3:  # check if user has entered something or not
            try:
                query = int(query)              # change ps no as integer
                searchid = "Ps No"              # ps no is saved as searchid
            except ValueError:
                if "@" in query:
                    searchid = "Email"
                else:
                    searchid = "Name"
            return query, searchid
        elif c == 3:
            print("Too many wrong attempts, try again later.")
            exit()
        else:
            print("No input found, try again!")
            return self.get_input(c + 1)

    # values of dataframe gets stored in x and we're updating the fields.

    def search(self, query, searchid):
        print(f"Searching {searchid.lower()} `{query}`...")   # searchid is converted to lowercase
        fields = {}
        for x in self.dfs.values():
            fields.update(
                x[x[searchid] == query].to_dict(orient="list"))  # checking if the query exist in that col or not
        # print(fields)

        if fields[searchid]:
            print("Found.")
            return pd.DataFrame.from_dict(fields)  # changing the dict into dataframe

        print(f"Couldn't find the {searchid.lower()} in sheets!")
        exit()

    def add_to_master(self, df):

        # df["Entry Time"] = df["Entry Time"].dt.strftime("%I:%S %p")
        # df["Exit Time"] = df["Exit Time"].dt.strftime("%H:%S %p")
        df["Start Date"] = df["Start Date"].dt.strftime("%d/%m/%Y")
        df["End Date"] = df["End Date"].dt.strftime("%d/%m/%Y")

        book = load_workbook(self.worksheet)  # loading the workbook
        with pd.ExcelWriter(self.worksheet, engine="openpyxl", mode="a") as writer:  # opening a Excel Writer instance
            writer.book = book  # changing the workbook of the writer to our current workbook , file is uploaded
            writer.sheets = {ws.title: ws for ws in book.worksheets}  # adding the worksheets to it
            # getting the last row if not found set it to 0
            try:
                startrow = writer.sheets[MSHEET].max_row
            except KeyError:
                startrow = 0
            # checking if we have already have a master sheet
            if MSHEET in writer.book.sheetnames:
                df.to_excel(
                    writer,
                    index=False,
                    header=False,
                    sheet_name=MSHEET,
                    startrow=startrow,
                )
            else:
                df.to_excel(writer, index=False, sheet_name=MSHEET, startrow=startrow)
        print(f"Added to {MSHEET}.")
        
        
        """
         This block is used to define the function to create the bar graph
         """
        
        def barchart(self):

        book = load_workbook(self.worksheet)
        chart1 = BarChart()
        chart1.type = "col"  # taken output of any two cols
        chart1.style = 10
        chart1.title = "Bar Chart"
        chart1.y_axis.title = 'Test number'
        chart1.x_axis.title = 'Sample length (mm)'

        data = Reference(book[MSHEET], min_col=1, min_row=4,
                         max_col=2, max_row=book[MSHEET].max_row)
        x = Reference(book[MSHEET], min_col=1, min_row=2, max_row=7)
        chart1.add_data(data, titles_from_data=True)
        chart1.set_categories(x)
        chart1.shape = 4
        book[MSHEET].add_chart(chart1, "A10")
        book.save(WORKSHEET)


no_of_inputs = int(input("Select the number of inputs: "))
for i in range(no_of_inputs):

    # main function

    if __name__ == "__main__":
        agg = Aggregator(WORKSHEET, SHEETS)

        query, searchid = agg.get_input()

        df_m = agg.search(query, searchid)

        agg.add_to_master(df_m)

        agg.BarChart()
